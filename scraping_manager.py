import os
import tablib
from datetime import datetime
from scrapers.facebook_scraper import FacebookScraper
from scrapers.youtube_scraper import YoutubeScraper
from scrapers.google_maps_scraper import GoogleMapsScraper
import logging
import openpyxl

class ScrapingManager:
    def __init__(self):
        self.setup_logging()
        self.facebook_scraper = FacebookScraper()
        self.youtube_scraper = YoutubeScraper()
        self.google_maps_scraper = GoogleMapsScraper()
        
    def setup_logging(self):
        """Setup logging configuration"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
    
    def load_targets(self, file_path='scraping_targets.xlsx'):
        """Load scraping targets from Excel file"""
        try:
            wb = openpyxl.load_workbook(file_path)
            
            targets = {
                'facebook': [],
                'youtube': [],
                'google_maps': []
            }
            
            # Load Facebook group URLs
            if 'Facebook' in wb.sheetnames:
                sheet = wb['Facebook']
                for row in sheet.iter_rows(min_row=2):
                    if row[0].value:  # URL column
                        targets['facebook'].append(row[0].value)
            
            # Load YouTube video URLs
            if 'YouTube' in wb.sheetnames:
                sheet = wb['YouTube']
                for row in sheet.iter_rows(min_row=2):
                    if row[0].value:  # URL column
                        targets['youtube'].append(row[0].value)
            
            # Load Google Maps search queries
            if 'GoogleMaps' in wb.sheetnames:
                sheet = wb['GoogleMaps']
                for row in sheet.iter_rows(min_row=2):
                    if row[0].value:  # Search query column
                        targets['google_maps'].append(row[0].value)
            
            return targets
            
        except Exception as e:
            self.logger.error(f"Error loading targets: {str(e)}")
            return None
    
    def scrape_all(self):
        """Run all scrapers and combine results"""
        all_leads = []
        targets = self.load_targets()
        
        if not targets:
            self.logger.error("No targets loaded. Please check scraping_targets.xlsx")
            return
        
        # Scrape Facebook groups
        if targets['facebook']:
            self.logger.info(f"Scraping {len(targets['facebook'])} Facebook groups...")
            facebook_leads = self.facebook_scraper.scrape(targets['facebook'])
            all_leads.extend(facebook_leads)
        
        # Scrape YouTube videos
        if targets['youtube']:
            self.logger.info(f"Scraping {len(targets['youtube'])} YouTube videos...")
            youtube_leads = self.youtube_scraper.scrape(targets['youtube'])
            all_leads.extend(youtube_leads)
        
        # Scrape Google Maps
        if targets['google_maps']:
            self.logger.info(f"Scraping {len(targets['google_maps'])} Google Maps queries...")
            maps_leads = self.google_maps_scraper.scrape(targets['google_maps'])
            all_leads.extend(maps_leads)
        
        return all_leads
    
    def export_leads(self, leads, format='xlsx'):
        """Export leads to file"""
        if not leads:
            self.logger.warning("No leads to export")
            return
        
        # Create dataset
        headers = ['name', 'email', 'phone', 'website', 'profile_url', 'content', 
                  'source_url', 'platform', 'type']
        
        dataset = tablib.Dataset(headers=headers)
        
        # Add data
        for lead in leads:
            row = [
                lead.get('name', ''),
                lead.get('email', ''),
                lead.get('phone', ''),
                lead.get('website', ''),
                lead.get('profile_url', ''),
                lead.get('content', ''),
                lead.get('source_url', ''),
                lead.get('platform', ''),
                lead.get('type', '')
            ]
            dataset.append(row)
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'leads_{timestamp}.{format}'
        
        # Export based on format
        if format == 'xlsx':
            with open(filename, 'wb') as f:
                f.write(dataset.export('xlsx'))
        elif format == 'csv':
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(dataset.export('csv'))
        
        self.logger.info(f"Exported {len(leads)} leads to {filename}")
        return filename

def main():
    manager = ScrapingManager()
    leads = manager.scrape_all()
    if leads:
        manager.export_leads(leads, 'xlsx')
        manager.export_leads(leads, 'csv')

if __name__ == "__main__":
    main()
